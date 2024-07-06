import random
import smtplib
from tkinter import *
import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import messagebox
import random,os,tempfile
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

import matplotlib.pyplot as plt

import pandas as pd

import sqlite3
def call():


    wb = Workbook()
    wb = load_workbook("C:/Users/jayan/Downloads/grocery new.xlsx")
    ws = wb.active

    def clear():
        meatentry.delete(0, END)
        for i in list1:
            meatentry.insert(0, 0)
        cannedentry.delete(0, END)
        for i in list1:
            cannedentry.insert(0, 0)
        diaryentry.delete(0, END)
        for i in list1:
            diaryentry.insert(0, 0)
        fruitentry.delete(0, END)
        for i in list1:
            fruitentry.insert(0, 0)
        snacksentry.delete(0, END)
        for i in list1:
            snacksentry.insert(0, 0)
        meatpriceentry.delete(0, END)
        diarypriceentry.delete(0, END)
        cannedpriceentry.delete(0, END)
        snackspriceentry.delete(0, END)
        fruitpriceentry.delete(0, END)

        nameentry.delete(0, END)
        phoneentry.delete(0, END)
        billnoentry.delete(0, END)

        text_widget.delete(0, END)

    def mail():
        def send():
            try:
                ob = smtplib.SMTP('smtp.gmail.com', 587)
                ob.starttls()
                ob.login(senderentry.get(), passwordentry.get())
                message = email_textarea.get(1.0, END)
                ob.sendmail(senderentry.get(), reciverentry.get(), message)
                ob.quit()
                messagebox.showinfo('Success', 'Bill is send successfully', parent=root1)
                root1.destroy()
            except:
                messagebox.showerror('Error', 'Something went wront please try again', parent=root1)

        if text_widget.get(1.0, END) == '\n':
            messagebox.showerror('Error', 'Bill is Empty')
        else:
            root1 = Toplevel()
            root1.grab_set()
            root1.title('Send E-Mail')
            root1.config(bg='gray20')
            root1.geometry('1000x1000')
            root1.resizable(0, 0)
            senderframe = LabelFrame(root1, text='SENDER', font=('arial', 16, 'bold'), bd=6, bg='gray20', fg='white',width=0,height=0)
            senderframe.grid(row=0, column=0, padx=80, pady=20)

            senderlable = Label(senderframe, text="SENDER'S E-MAIL", font=('arial', 14, 'bold'), bd=6, bg='gray20',
                                fg='white')
            senderlable.grid(row=0, column=0, pady=8, padx=10)

            senderentry = Entry(senderframe, font=('arial', 14, 'bold'), bd=2, width=23, relief=RIDGE)
            senderentry.grid(row=1, column=0, padx=100, pady=30)

            senderlable = Label(senderframe, text="PASSWORD", font=('arial', 14, 'bold'), bd=6, bg='gray20',
                                fg='white')
            senderlable.grid(row=0, column=1, pady=8, padx=10)

            passwordentry = Entry(senderframe, font=('arial', 14, 'bold'), bd=2, width=23, relief=RIDGE, show='*')
            passwordentry.grid(row=1, column=1, padx=10, pady=8)

            reciverframe = LabelFrame(root1, text='RECIPIENT', font=('arial', 16, 'bold'), bd=6, bg='gray20',
                                      fg='white',width=400,height=400)
            reciverframe.grid(row=5, column=0, padx=40, pady=20)

            reciverlable = Label(reciverframe, text="E-MAIL ADDERSS", font=('arial', 14, 'bold'), bd=6, bg='gray20',
                                 fg='white')
            reciverlable.grid(row=0, column=0, pady=60, padx=100)

            reciverentry = Entry(reciverframe, font=('arial', 14, 'bold'), bd=2, width=23, relief=RIDGE)
            reciverentry.grid(row=0, column=1, padx=10, pady=8)

            messagelable = Label(reciverframe, text="MESSAGE", font=('arial', 14, 'bold'), bd=6, bg='gray20',
                                 fg='white')
            messagelable.grid(row=1, column=0, pady=8, padx=10)

            email_textarea = Text(reciverframe, font=('arial', 14, 'bold'), bd=2, relief=SUNKEN, width=42, height=11)
            email_textarea.grid(row=1, column=0, columnspan=2)
            email_textarea.delete(1.0, END)
            email_textarea.insert(END,
                                  text_widget.get(1.0, END).replace('=', '').replace('-', '').replace('\t\t\t', '\t\t'))

            sendbutton = Button(root1, text='SEND', font=('arial', 16, 'bold'), width=15, command=send)
            sendbutton.grid(row=4, column=0, pady=20)

            root1.mainloop()

    def print_bill():
        if text_widget.get(1.0, END) == '\n':
            messagebox.showerror('Error', 'Bill is Empty')
        else:
            file = tempfile.mktemp(' .txt')
            open(file, 'w').write(text_widget.get(1.0, END))
            os.startfile(file, 'print')

    def search_bill():
        for i in os.listdir('bills/'):
            if i.split('.')[0] == billnoentry.get():
                #f = open(f'bills/{i}', 'r')
                text_widget.delete(1.0, END)
                with open(f'bills/{i}', 'r', encoding='latin-1') as f:
                    for data in f:
                        text_widget.insert(END, data)
                f.close()
                break
            else:
                messagebox.showerror('Error', 'Invalid Bill Number')


            # Process the data

    if not os.path.exists('bills'):
        os.mkdir('bills')

    def save():
        global billno

        result = messagebox.askyesno('Confirm', 'Do you want to save the bill?')
        if result:
            bill_content = text_widget.get(1.0, END)
            file = open(f'bills/ {billno}.txt', 'w')
            file.write(bill_content)
            file.close()
            messagebox.showinfo('Success', f'{billno} is saved successfully')
            billno = random.randint(500, 1000)

    billno = random.randint(500, 1000)

    def bill():
        if nameentry.get() == '' or phoneentry.get() == '':
            messagebox.showerror('Error', 'Customer details are required')
        elif meatpriceentry.get() == '' and diarypriceentry.get() == '' and cannedpriceentry.get() == '' and fruitpriceentry.get() == '' and snackspriceentry.get() == '':
            messagebox.showerror('Error', 'No products are selected')
        elif meatpriceentry.get() == '0 Rs' and diarypriceentry.get() == '0 Rs' and cannedpriceentry.get() == '0 Rs' and fruitpriceentry.get() == '0 Rs' and snackspriceentry.get() == '0 Rs':
            messagebox.showerror('Error', 'No products are selected')
        else:
            text_widget.delete(1.0, END)

            text_widget.insert(END, '\t\t**Welcome Customer**\n')
            text_widget.insert(END, f'\nBill Number: {billno}\n')
            text_widget.insert(END, f'\nCustomer Name: {nameentry.get()}\n')
            text_widget.insert(END, f'\nCustomer Phone Number: {phoneentry.get()}\n')
            text_widget.insert(END, '\n==============================\n')
            text_widget.insert(END, 'Product\t\t\tQuantity\t\t\tPrice')
            text_widget.insert(END, '\n==============================')
            for i in range(10):
                if l[i] == 0:
                    continue
                else:
                    text_widget.insert(END, f'\n{product1[i].value}\t\t\t{l[i]}\t\t\t{totprice1} Rs')

            for i in range(10):
                if l1[i] == 0:
                    continue
                else:
                    text_widget.insert(END, f'\n{product2[i].value}\t\t\t{l1[i]}\t\t\t{totprice2} Rs')

            for i in range(10):
                if l2[i] == 0:
                    continue
                else:
                    text_widget.insert(END, f'\n{product3[i].value}\t\t\t{l2[i]}\t\t\t{totprice3} Rs')

            for i in range(10):
                if l3[i] == 0:
                    continue
                else:

                    text_widget.insert(END, f'\n{product4[i].value}\t\t\t{l3[i]}\t\t\t{totprice4} Rs')

            for i in range(10):
                if l4[i] == 0:
                    continue
                else:
                    text_widget.insert(END, f'\n{product5[i].value}\t\t\t{l4[i]}\t\t\t{totprice5} Rs')

    def tot():
        global l, l1, l2, l3, l4, p1, p2, p3, p4, p5, totprice1, totprice2, totprice3, totprice4, totprice5
        p1 = ws['L']
        meatquan = []
        price = []
        # meatlist.clear()

        l = []
        for i in meatlist:
            l.append(int(i.get()))

        for i in p1:
            meatquan.append(i.value)
        print(meatquan)

        for i in range(len(l)):
            price.append(int(l[i]) * meatquan[i])

        totprice1 = sum(price)
        print(totprice1)

        meatpriceentry.insert(0, str(totprice1) + ' Rs')

        # meatpriceentry.delete(0,END)

        p2 = ws['C']
        diaryquan = []
        price1 = []
        for i in p2:
            diaryquan.append(i.value)
        l1 = []
        for i in diarylist:
            l1.append(int(i.get()))
        for i in range(len(diarylist)):
            price1.append(l1[i] * diaryquan[i])

        totprice2 = sum(price1)
        diarypriceentry.insert(0, str(totprice2) + ' Rs')
        price1.clear()
        # diarypriceentry.delete(0, END)

        p3 = ws['F']
        cannedquan = []
        price2 = []
        for i in p3:
            cannedquan.append(i.value)
        l2 = []
        for i in cannedlist:
            l2.append(int(i.get()))

        for i in range(len(cannedlist)):
            price2.append(int(l2[i]) * cannedquan[i])

        totprice3 = sum(price2)
        cannedpriceentry.insert(0, str(totprice3) + ' Rs')
        price2.clear()
        # cannedpriceentry.delete(0, END)

        p4 = ws['I']
        fruitquan = []
        price3 = []
        for i in p4:
            fruitquan.append(i.value)
        l3 = []
        for i in fruitlist:
            l3.append(int(i.get()))

        for i in range(len(fruitlist)):
            price3.append(int(l3[i]) * fruitquan[i])

        totprice4 = sum(price3)
        fruitpriceentry.insert(0, str(totprice4) + ' Rs')
        price3.clear()
        # fruitpriceentry.delete(0, END)

        p5 = ws['O']
        snacksquan = []
        price4 = []
        for i in p5:
            snacksquan.append(i.value)
        l4 = []
        for i in snackslist:
            l4.append(int(i.get()))
        print(l)

        for i in range(len(snackslist) - 1):
            price4.append(int(l4[i]) * snacksquan[i])

        totprice5 = sum(price4)
        snackspriceentry.insert(0, str(totprice5) + 'Rs')
        price4.clear()
        # snackspriceentry.delete(0, END)

    root = Tk()
    root.title('grocery billing system')
    root.geometry('1500x1000')
    heading = Label(root, text='billing system', font=('times new roman', 30, 'bold'), bg='gray20', fg='orange', bd=12,
                    relief=GROOVE).pack(fill=X)

    customer_details_frame = LabelFrame(root, text='CUSTOMER DETAILS', font=('times new roman', 15, 'bold'),
                                        bg='gray20',
                                        fg='gold', bd=8, relief=GROOVE).pack()

    namelabel = Label(customer_details_frame, text="NAME", font=('times new roman', 15, 'bold'), bg='gray20',
                      fg='orange').place(x=5, y=78)
    nameentry = Entry(customer_details_frame, font=('arial', 15), bd=7, width=18)
    nameentry.place(y=75, x=90)

    phonelable = Label(customer_details_frame, text="PHONE", font=('times new roman', 15, 'bold'), bg='gray20',
                       fg='orange').place(x=350, y=78)
    phoneentry = Entry(customer_details_frame, font=('arial', 15), bd=7, width=18)
    phoneentry.place(y=75, x=450)

    maillabel = Label(customer_details_frame, text="E-MAIL", font=('times new roman', 15, 'bold'), bg='gray20',
                      fg='orange').place(x=700, y=78)
    mailentey = Entry(customer_details_frame, font=('arial', 15), bd=7, width=18)
    mailentey.place(y=75, x=800)

    billnolable = Label(customer_details_frame, text='BILL NO', font=('times new roman', 15, 'bold'), bg='gray20',
                        fg='orange').place(x=1050, y=78)
    billnoentry = Entry(customer_details_frame, font=('arial', 15), bd=7, width=18)
    billnoentry.place(y=75, x=1160)

    searchbutton = Button(customer_details_frame, text='SEARCH', font=('arial', 12, 'bold'), bd=7, width=10,
                          command=search_bill)
    searchbutton.place(x=1380, y=73)
    '''
    diaryframe = Frame(root,bg='gray20',width=250,height=450,bd=8,relief=GROOVE).place(x=0,y=130)
    
    fruitframe = Frame(root,bg='gray20',width=250,height=450,bd=8,relief=GROOVE).place(x=251,y=130)
    
    meatframe = Frame(root,bg='gray20',width=250,height=450,bd=8,relief=GROOVE).place(x=502,y=130)
    
    snacksframe = Frame(root,bg='gray20',width=250,height=450,bd=8,relief=GROOVE).place(x=753,y=130)
    
    juiceframe = Frame(root,bg='gray20',width=250,height=450,bd=8,relief=GROOVE).place(x=1004,y=130)
    
    diarylable = Label(diaryframe,text='DIARY',font=('times new roman',15,'bold'),bg='gray20',fg='orange').place(x=90,y=135)
    
    fruitlable = Label(fruitframe,text='FRUIT',font=('times new roman',15,'bold'),bg='gray20',fg='orange').place(x=340,y=135)
    
    meatlable = Label(meatframe,text='MEAT',font=('times new roman',15,'bold'),bg='gray20',fg='orange').place(x=600,y=135)
    
    snackslable = Label(snacksframe,text='SNACKS',font=('times new roman',15,'bold'),bg='gray20',fg='orange').place(x=840,y=135)
    
    juicelable = Label(juiceframe,text='JUICE',font=('times new roman',15,'bold'),bg='gray20',fg='orange').place(x=1100,y=135)
    
    
    
    #def create_frame_with_scrollbar(parent,width, height):
        global meatlist
        global cannedlist
        global diarylist
        global fruitlist
        global snackslist
    
    '''
    parent = root
    width = 200
    height = 450
    product1 = ws['J']
    list1 = []
    for i in product1:
        list1.append(i.value)

    frame = tk.Frame(parent, bd=8, bg='gray20', width=width, height=height, relief=GROOVE)
    frame.place(x=0, y=130)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas = tk.Canvas(frame, width=width, height=height, yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=canvas.yview)

    inner_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor=tk.NW)

    # Add labels and entries to the inner frame
    meatlist = []
    for i in list1:
        label = tk.Label(inner_frame, text=f"{i}")
        label.pack(pady=20)
        meatentry = tk.Entry(inner_frame)
        meatentry.pack(pady=1)
        meatentry.insert(0, 0)
        meatlist.append(meatentry)

    # Configure canvas scroll region
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    product2 = ws['D']
    list1 = []
    for i in product2:
        list1.append(i.value)

    frame = tk.Frame(parent, width=width, height=height, bd=8, bg='gray20', relief=GROOVE)
    frame.place(x=235, y=130)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas = tk.Canvas(frame, width=width, height=height, yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=canvas.yview)

    inner_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor=tk.NW)

    # Add labels and entries to the inner frame
    cannedlist = []
    for i in list1:
        label = tk.Label(inner_frame, text=f"{i}")
        label.pack(pady=20)
        cannedentry = tk.Entry(inner_frame)
        cannedentry.pack(pady=1)
        cannedentry.insert(0, 0)
        cannedlist.append(cannedentry)

    # Configure canvas scroll region
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    product3 = ws['A']
    list1 = []
    for i in product3:
        list1.append(i.value)

    frame = tk.Frame(parent, width=width, height=height, bd=8, bg='gray20', relief=GROOVE)
    frame.place(x=470, y=130)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas = tk.Canvas(frame, width=width, height=height, yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=canvas.yview)

    inner_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor=tk.NW)

    # Add labels and entries to the inner frame
    diarylist = []
    for i in list1:
        label = tk.Label(inner_frame, text=f"{i}")
        label.pack(pady=20)
        diaryentry = tk.Entry(inner_frame)
        diaryentry.pack(pady=1)
        diaryentry.insert(0, 0)
        diarylist.append(diaryentry)

    # Configure canvas scroll region
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    product4 = ws['G']
    list1 = []
    for i in product4:
        list1.append(i.value)

    frame = tk.Frame(parent, width=width, height=height, bd=8, bg='gray20', relief=GROOVE)
    frame.place(x=705, y=130)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas = tk.Canvas(frame, width=width, height=height, yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=canvas.yview)

    inner_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor=tk.NW)

    # Add labels and entries to the inner frame
    fruitlist = []
    for i in list1:
        label = tk.Label(inner_frame, text=f"{i}")
        label.pack(pady=20)
        fruitentry = tk.Entry(inner_frame)
        fruitentry.pack(pady=1)
        fruitentry.insert(0, 0)
        fruitlist.append(fruitentry)
    # Configure canvas scroll region
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    product5 = ws['M']
    list1 = []
    for i in product5:
        list1.append(i.value)

    frame = tk.Frame(parent, width=width, height=height, bd=8, bg='gray20', relief=GROOVE)
    frame.place(x=940, y=130)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas = tk.Canvas(frame, width=width, height=height, yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=canvas.yview)

    inner_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor=tk.NW)

    # Add labels and entries to the inner frame
    snackslist = []
    for i in list1:
        label = tk.Label(inner_frame, text=f"{i}")
        label.pack(pady=20)
        snacksentry = tk.Entry(inner_frame)
        snacksentry.pack(pady=1)
        snacksentry.insert(0, 0)

        snackslist.append(snacksentry)

    # Configure canvas scroll region
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    # Create five frames with scrollbars

    ############create_frame_with_scrollbar(root, width=200, height=450)
    # frame1 = Frame(root, width=350, height=400, bd=8, relief=GROOVE)
    # frame1.place(x=1180,y=200)

    # Create a frame
    frame = tk.Frame(root, bd=8)
    frame.place(x=1180, y=200)

    # Create a Text widget
    text_widget = tk.Text(frame, width=40, height=24)
    text_widget.pack(side=tk.LEFT)

    # Create a Scrollbar
    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Configure the Text widget to use the Scrollbar
    text_widget.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=text_widget.yview)

    priceframe = Frame(root, width=700, height=185, bd=8, bg='gray20', relief=GROOVE).place(x=0, y=600)

    meatpricelabel = Label(priceframe, text="MEAT", font=('times new roman', 15, 'bold'), bg='gray20',
                           fg='orange').place(x=10, y=625)
    meatpriceentry = tk.Entry(priceframe, font=('arial', 15), bd=7, width=10)
    meatpriceentry.place(x=100, y=625)

    cannedpricelabel = Label(priceframe, text="CANNED", font=('times new roman', 15, 'bold'), bg='gray20',
                             fg='orange').place(x=10, y=700)
    cannedpriceentry = tk.Entry(priceframe, font=('arial', 15), bd=7, width=10)
    cannedpriceentry.place(y=700, x=100)

    diarypricelabel = Label(priceframe, text="DIARY", font=('times new roman', 15, 'bold'), bg='gray20',
                            fg='orange').place(x=240, y=625)
    diarypriceentry = tk.Entry(priceframe, font=('arial', 15), bd=7, width=10)
    diarypriceentry.place(y=625, x=320)

    fruitpriceabel = Label(priceframe, text="FRUITS", font=('times new roman', 15, 'bold'), bg='gray20',
                           fg='orange').place(x=240, y=700)
    fruitpriceentry = tk.Entry(priceframe, font=('arial', 15), bd=7, width=10)
    fruitpriceentry.place(y=700, x=320)

    snackspricelabel = Label(priceframe, text="SNACKS", font=('times new roman', 15, 'bold'), bg='gray20',
                             fg='orange').place(x=460, y=625)
    snackspriceentry = tk.Entry(priceframe, font=('arial', 15), bd=7, width=10)
    snackspriceentry.place(y=625, x=565)

    totbutton = Button(text="TOTAL", font=('arial', 16, 'bold'), bg='gray20', fg='white', bd=5, width=8, command=tot)
    totbutton.place(x=750, y=680)
    billbutton = Button(text="BILL", font=('arial', 16, 'bold'), bg='gray20', fg='white', bd=5, width=8,
                        command=bill).place(x=900, y=680)
    emailbutton = Button(text="E-MAIL", font=('arial', 16, 'bold'), bg='gray20', fg='white', bd=5, width=8,
                         command=mail).place(x=1050, y=680)
    printbutton = Button(text="PRINT", font=('arial', 16, 'bold'), bg='gray20', fg='white', bd=5, width=8,
                         command=print_bill).place(x=1200, y=680)
    clearbutton = Button(text="CLEAR", font=('arial', 16, 'bold'), bg='gray20', fg='white', bd=5, width=8,command=clear).place(x=1350,y=680)



    def analyze_data():
        global lable1,lable2

        # Create the main window
        window = tk.Tk()
        window.title("Scrollable Window")

        # Create a frame to hold the scrollbar and the content
        frame = ttk.Frame(window)
        frame.pack(fill=tk.BOTH, expand=True)

        # Create a scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Create a canvas
        canvas = tk.Canvas(frame, yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Configure the scrollbar to scroll the canvas
        scrollbar.config(command=canvas.yview)

        # Create another frame inside the canvas
        content_frame = ttk.Frame(canvas)

        # Add some widgets to the content frame (e.g., labels, buttons, etc.0)

        for i in product1:

            label1 = tk.Label(window, text=f"{i.value}")
            #label1.pack(padx=10,pady=10)
            label1.grid(row=0,column=0)
            # label = tk.Label(window, text=f"Label {i}")
            # label.pack(padx=10, pady=10)

        for i in p1:
            label2 = tk.Label(window, text=f"{i.value}")
            #label2.pack(padx=10, pady=10)
            label2.grid(row=0, column=1)
        # Add the content frame to the canvas
        canvas.create_window(0, 0, anchor=tk.NW, window=content_frame)

        # Configure the canvas to update scroll region when the size of the content frame changes
        content_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))

        # Run the Tkinter event loop
        window.mainloop()




        def graph(product1, product2, product3, product4, product5, p1, p2, p3, p4, p5):
            plt.plot(product1, p1, lable='MEAT', marker='o', markerfacecolor='green')
            plt.plot(product2, p2, lable='CANNED_PRODUCT', marker='o', markerfacecolor='yellow')
            plt.plot(product3, p3, lable='DIARY', marker='o', markerfacecolor='violet')
            plt.plot(product4, p4, lable='FRUIT', marker='o', markerfacecolor='blue')

            plt.plot(product5, p5, lable='SNACKS', marker='o', markerfacecolor='red')
            plt.xlabel('PRODUCTS')
            plt.ylabel('SALES')
            plt.title('SALES GRAPH')
            plt.legend()
            plt.show()



    anabutton = Button(text="ANALIZE", font=('arial', 16, 'bold'), bg='gray20', fg='white', bd=5, width=8,command="graph").place(x=1350,
                                                                                                                 y=620)



    root.mainloop()













