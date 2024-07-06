import random
import smtplib
from tkinter import *
import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import messagebox
import random,os,tempfile
from billing import call
from PIL import Image
import PIL.Image

import sqlite3

root2 = Tk()
root2.title("Login")
root2.geometry("825x850")
root2.configure(bg="white")
root2.resizable(False, False)

Fullname=StringVar()
Email=StringVar()
var=IntVar()
c=StringVar()
var1=IntVar()

wb = Workbook()
wb = load_workbook("C:/Users/jayan/Downloads/grocery new.xlsx")
ws = wb.active

list1 = []
list3 = []
list2 = []

pri_list = []




img = PhotoImage(file="C:/Users/jayan/OneDrive/Documents/python/pythonProject2/gui/clipart-man-grocery-shopping-6.png")
Label(root2, image=img, bg="white").place(x=50, y=0)



##################----------------------------------------------------------
label_1= Label(root2,text="FullName",width=20,font=("bold",10))
label_1.place(x=200,y=450)
entry_1 =Entry(root2,textvar=Fullname)
entry_1.place(x=450,y=450)
global user


label_2= Label(root2,text="password",width=20,font=("bold",10))
label_2.place(x=200,y=500)

entry_2=Entry(root2,textvar=Email)
entry_2.place(x=450,y=500)







def mydelete():
    entry_1.delete(0,END)
    entry_2.delete(0,END)




##################----------------------------------------------------------#

# code = Entry(frame,width = 25, fg="black",bg = "white", border = 0,font = ("Microsoft YaHei Light",11,"bold"))
# code.place(x = 30,y = 150)
# code.insert(0,"Password:")

# Frame(frame,width = 200,height = 2, bg = "black").place(x = 25,y = 107)
def ok():
    print("ok")

def data():
    pass


def mysubmit():
    root2.destroy()
    call()


frame8 = Frame(root2, width=50, height=10, bg="white")
frame8.place(x=280, y=620)

my_button = Button(frame8, text="SUBMIT", font="Helvetica", command=mysubmit)
my_button.pack(padx=0, pady=10)

frame9 = Frame(root2, width=50, height=10, bg="white")
frame9.place(x=500, y=620)
b_button = Button(frame9, text="CLEAR", font="Helvetica", command=mydelete).pack(padx=0, pady=10)



root2.mainloop()